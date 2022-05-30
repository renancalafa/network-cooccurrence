from ftplib import all_errors
import pandas as pd
import numpy as np
import networkx as nx
import os
import openpyxl as op
from openpyxl.styles import Border, Side, Font

class Comparator:

    comorbidities = ['cancer', 'renalfailure', 'pulmonarydisease', 'cardiacdisease', 'obesity', 'pregnancy', 'smoke', 'sot', 'diabetes', 'cbvsc', 'hypertension', 'hivaids', 'liverdisease', 'neuro ', 'paralysis', 'hypothyroid', 'rheum', 'coag', 'danemia ', 'dabuse', 'psychoses', 'depression']

    def create_sheets(self, races):
        wb = op.Workbook()

        races_info = self.set_races_data(races)
        self.create_main_sheet(wb, races_info)
        self.create_top10_sheet(wb, races_info)

        wb.save("sheets/Main.xlsx") 


    def create_main_sheet(self, wb, races):
        ws = wb.active
        ws.title = "All Cooccurrences"
        for i, race in enumerate(races):
            self.set_main_worksheet(ws, race)   

            race_dict_sorted = sorted(race["race_data_all"], key=lambda d: d['Phi'], reverse=True) 

            for row, data in enumerate(race_dict_sorted, start=3):
                ws[f"B{row}"] = data["Source"]
                ws[f"C{row}"] = data["Target"]
                ws[f"D{row}"] = data["Phi"]
                ws[f"E{row}"] = data["t"]

            ws.move_range("A1:E233", cols = 5 * (3-i))        

        self.highlight_not_sig(ws)

    def create_top10_sheet(self, wb, races):

        ws = wb.create_sheet("Top 10 Cooccurrences")

        cmbdts_sorted_lists = []                
        for race in races:
            cmbdts_indexed_lists = []
            for i, obj in enumerate(race["race_data_all"]):
                obj["cc_index"] = i
                cmbdts_indexed_lists.append(obj)
            cmbdts_sorted_lists.append(sorted(cmbdts_indexed_lists, key=lambda d: d['Phi'], reverse=True)[0:20])

        cc_list = self.set_top10_ccs(cmbdts_sorted_lists)

        self.set_top10_worksheet(ws, races, cc_list)

    def set_top10_ccs(self, lists):
        cmbdts = []
        for j in range(0,20):
            for i in range(0,4):
                if not any((d['Source'] == lists[i][j]["Source"]) & (d["Target"] == lists[i][j]["Target"]) for d in cmbdts):
                    cmbdts.append(lists[i][j])
                    if (len(cmbdts) > 12):
                        return cmbdts            

    def set_races_data (self, races):
        races_data_list = races
        for i, race in enumerate(races):
            list_dict = self.load_data(race)
            for value in list_dict:
                race["race_data" + value["sufix"]] = value["data"]
            
        return races_data_list  

    def load_data(self, race):
        all_data = []
        composite_sufixes = ['_all', '_c', '_nc']
        for sufix in composite_sufixes:
            phi_df = pd.read_csv('test/' + race["race_abv"] + '/' + race["race_abv"] + sufix + '_Phi.csv', header=None)
            t_df = pd.read_csv('test/' + race["race_abv"] + '/' + race["race_abv"] + sufix + '_t.csv', header=None)
            cc_df = pd.read_csv('test/' + race["race_abv"] + '/' + race["race_abv"] + sufix + '_CC.csv', header=None)
            all_data.append(self.set_race_dict(phi_df, t_df, cc_df, sufix))
        return all_data

    def set_top10_worksheet(self, ws, races, cc_list):

        title_cells_range = ['C1:H1', 'J1:O1', 'Q1:V1', 'X1:AC1']
        for range in title_cells_range:
            ws.merge_cells(range)
            ws[range.split(":",1)[0]].alignment = op.styles.Alignment(horizontal='center', vertical='center')

        composite_cells_range = ['C2:E2', 'F2:H2', 'J2:L2', 'M2:O2', 'Q2:S2', 'T2:V2', 'X2:Z2', 'AA2:AC2']
        for range in composite_cells_range:
            ws.merge_cells(range)
        #     ws[range.split(":",1)[0]].alignment = op.styles.Alignment(horizontal='center', vertical='center')
        #     ws[range.split(":",1)[0]].font = Font(bold=True)

        # bd = Border(top = Side(border_style='thin', color='FF000000'),    
        #                       right = Side(border_style='thin', color='FF000000'), 
        #                       bottom = Side(border_style='thin', color='FF000000'),
        #                       left = Side(border_style='thin', color='FF000000')) 

      
        colums_low_width = ['I', 'P', 'W']
        for column in colums_low_width:
            ws.column_dimensions[column].width = 1

        for k, cc in enumerate(cc_list):
            ws["A" + str(k+4)].value = cc["Source"]
            ws["B" + str(k+4)].value = cc["Target"]


        for i, race in enumerate(races):
            ws["C1"].value = race["race_full"]
            ws["C1"].font = Font(bold=True)
            ws["C1"].alignment = op.styles.Alignment(horizontal='center', vertical='center')

            cells_type_align = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            for cell in cells_type_align:
                ws[cell + '3'].alignment = op.styles.Alignment(horizontal='center', vertical='center')
                ws[cell + '3'].font = Font(bold=True)
            ws["A3"].value = "Source"
            ws["B3"].value = "Target"
            ws["C3"].value = "N"
            ws["D3"].value = "Phi"
            ws["E3"].value = "t"
            ws["F3"].value = "N"
            ws["G3"].value = "Phi"
            ws["H3"].value = "t"


            # cells_bold = ['C2','F2']
            # for i, cell in enumerate(cells_bold):
            #     ws[cell].value = 0  if i == 0 else 1
            #     ws[cell].alignment = op.styles.Alignment(horizontal='center', vertical='center')
            ws["C2"].value = 0
            ws["C2"].alignment = op.styles.Alignment(horizontal='center', vertical='center')
            ws["C2"].font = Font(bold=True)
            ws["F2"].value = 1
            ws["F2"].alignment = op.styles.Alignment(horizontal='center', vertical='center')
            ws["F2"].font = Font(bold=True)

            for j, cc in enumerate(cc_list):
                ws["C" + str(j+4)].value = race["race_data_nc"][cc["cc_index"]]["N"]
                ws["D" + str(j+4)].value = race["race_data_nc"][cc["cc_index"]]["Phi"]
                ws["E" + str(j+4)].value = race["race_data_nc"][cc["cc_index"]]["t"]
                ws["F" + str(j+4)].value = race["race_data_c"][cc["cc_index"]]["N"]
                ws["G" + str(j+4)].value = race["race_data_c"][cc["cc_index"]]["Phi"]
                ws["H" + str(j+4)].value = race["race_data_c"][cc["cc_index"]]["t"]
                if(-1.96 < race["race_data_nc"][cc["cc_index"]]["t"] < 1.96):
                 ws["E" + str(j+4)].font = Font(bold=True)
                 ws["D" + str(j+4)].font = Font(bold=True)
                if(-1.96 < race["race_data_c"][cc["cc_index"]]["t"] < 1.96):
                 ws["G" + str(j+4)].font = Font(bold=True)
                 ws["H" + str(j+4)].font = Font(bold=True)

            ws.move_range("C1:H16", cols = 7 * (3-i))

        cells_color_ranges = ['C1:H16', 'J1:O16', 'Q1:V16', 'X1:AC16']
        race_colours = ['hisp', 'white', 'asian', 'black']
        for l, range in enumerate(cells_color_ranges):
            for row in ws[range]:
                for cell in row:
                    cell.fill = self.get_cell_colour(race_colours[l])

    def set_main_worksheet(self, ws, race):

        colour = self.get_cell_colour(race["race_abv"])
        ws.cell(row=1,column=2).value = race["race_full"]
        ws.cell(row=1,column=2).alignment = op.styles.Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B1:E1')
        ws.merge_cells('G1:J1')
        ws.merge_cells('L1:O1')
        ws.merge_cells('Q1:T1')
        ws["B1"].fill = colour
        ws["B2"] = "Source"
        ws["C2"] = "Target"
        ws["D2"] = "Phi"
        ws["E2"] = "t"
        ws.column_dimensions['F'].width = 1
        ws.column_dimensions['K'].width = 1
        ws.column_dimensions['P'].width = 1

        bd = Border(top = Side(border_style='thin', color='FF000000'),    
                              right = Side(border_style='thin', color='FF000000'), 
                              bottom = Side(border_style='thin', color='FF000000'),
                              left = Side(border_style='thin', color='FF000000')) 

        for row in ws["B1:E233"]:
            for cell in row:
                cell.fill = colour
                cell.border = bd

    def set_race_dict(self, phi_df, t_df, cc_df, sufix):
        cmbdts = self.comorbidities
        race_dict_list = []
        phi_dict = phi_df.to_dict('records')
        t_dict = t_df.to_dict('records')
        cc_dict = cc_df.to_dict('records')
        for i, data in enumerate(phi_dict):
            for j, key in enumerate(data):
                if(j > i): #eliminates equals and duplicated coocurrences
                    race_dict_list.append({
                        "Source": cmbdts[i],
                        "Target": cmbdts[key],
                        "Phi": data[key],
                        "t": t_dict[i][j],
                        "N": cc_dict[i][j], 
                    })
        race_dict = {
            "sufix": sufix,
            "data": race_dict_list
        }
        return race_dict

    def get_cell_colour(self, race):
        if(race == "black"):
            colour = op.styles.PatternFill(start_color="f4cccc", fill_type="solid")
        elif(race == "asian"):
            colour = op.styles.PatternFill(start_color="d9ead3", fill_type="solid")
        elif(race == "white"):
            colour = op.styles.PatternFill(start_color="d9d2e9", fill_type="solid")
        elif(race == "hisp"):
            colour = op.styles.PatternFill(start_color="c9daf8", fill_type="solid")
        else:
            colour = op.styles.PatternFill(start_color="fce5cd", fill_type="solid")
        return colour

    def highlight_not_sig (self, ws):

        range_cols = ["D3:E233", 'I3:J233', 'N3:O233', 'S3:T233']
        for i, range_col in enumerate(range_cols):
            for row in ws[range_col]:
                if(-1.96 < row[1].value < 1.96):
                    for cell in row:
                        cell.font = Font(bold=True)